import os
import sqlite3
import zipfile
from datetime import datetime
from io import BytesIO
import re

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from PIL import Image, UnidentifiedImageError
from reportlab.pdfgen import canvas
import streamlit as st
import requests
import PIL

# Streamlit page configuration
st.set_page_config(
    page_title="Expense Report Processor",
    initial_sidebar_state="expanded",
    page_icon="📊",
    layout="wide"
)

# Initialize session state for history
if "processing_history" not in st.session_state:
    st.session_state.processing_history = []

def init_database():
    conn = sqlite3.connect('expense_processor.db')
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS processing_history
        (timestamp TEXT, filename TEXT, processed_file TEXT, 
         total_receipts INTEGER, processed_receipts INTEGER)
    ''')
    conn.commit()
    conn.close()

def add_to_history(data):
    conn = sqlite3.connect('expense_processor.db')
    c = conn.cursor()
    c.execute('''
        INSERT INTO processing_history
        (timestamp, filename, processed_file, total_receipts, processed_receipts)
        VALUES (?, ?, ?, ?, ?)
    ''', (
        data["timestamp"], data["filename"], data["processed_file"],
        data["total_receipts"], data["processed_receipts"]
    ))
    conn.commit()
    conn.close()

def get_processing_history():
    conn = sqlite3.connect('expense_processor.db')
    df = pd.read_sql_query(
        'SELECT * FROM processing_history ORDER BY timestamp DESC', 
        conn
    )
    conn.close()
    return df.to_dict('records')

def get_output_filename(input_filename):
    # Extract date from input filename (e.g., expenses_2025-1.csv -> 2025-1)
    match = re.search(r'expenses_(\d{4}-\d+)\.csv', input_filename)
    period = match.group(1) if match else datetime.now().strftime("%Y-%m")
    return f"processed_expenses_{period}"

def format_date(date_string):
    formats = ["%d. %B %Y", "%d.%m.%Y", "%d.%m.%y"]
    for fmt in formats:
        try:
            return datetime.strptime(date_string, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return date_string

def create_excel_file(df, filename):
    excel_path = f"{filename}.xlsx"
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        worksheet = writer.book["Sheet1"]

        # Style header
        for cell in worksheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Adjust column widths
        for col in worksheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            worksheet.column_dimensions[col[0].column_letter].width = max_length + 2

        # Add conditional formatting for receipt status
        worksheet.conditional_formatting.add(
            f'K2:K{worksheet.max_row}',
            Rule(
                type="containsText",
                operator="containsText",
                text="Available",
                dxf=DifferentialStyle(fill=PatternFill(bgColor="C6EFCE"))
            )
        )
        worksheet.conditional_formatting.add(
            f'K2:K{worksheet.max_row}',
            Rule(
                type="containsText",
                operator="containsText",
                text="Missing",
                dxf=DifferentialStyle(fill=PatternFill(bgColor="FFC7CE"))
            )
        )
        worksheet.conditional_formatting.add(
            f'K2:K{worksheet.max_row}',
            Rule(
                type="containsText",
                operator="containsText",
                text="Failed",
                dxf=DifferentialStyle(fill=PatternFill(bgColor="FFEB9C"))
            )
        )

        # Alternate row colors
        for row in range(2, worksheet.max_row + 1):
            if row % 2 == 0:
                for cell in worksheet[row]:
                    cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    return excel_path

def create_pdf_from_image(image_url):
    if not image_url or not isinstance(image_url, str):
        return None
        
    try:
        # Set up headers to mimic a browser request
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Accept': 'image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.9',
            'Referer': 'https://spesr.natron.io/',
            'Connection': 'keep-alive',
        }
        
        # Get the image data from the URL with headers
        response = requests.get(image_url, headers=headers, timeout=15, allow_redirects=True)
        response.raise_for_status()
        
        # Debug information
        content_length = len(response.content)
        content_type = response.headers.get('Content-Type', '')
        st.write(f"Response status: {response.status_code}, Content-Type: {content_type}, Content length: {content_length}")
        
        # Check if we received any content
        if not response.content:
            st.error(f"Empty response received from URL: {image_url}")
            return None
            
        img_data = BytesIO(response.content)
        
        # Check if the BytesIO object has content
        if img_data.getbuffer().nbytes == 0:
            st.error(f"Zero bytes received from URL: {image_url}")
            return None
        
        # Reset the buffer position to the beginning
        img_data.seek(0)
        
        # Try to open the image and convert to RGB mode to ensure compatibility
        img = Image.open(img_data)
        
        # For large images, resize to a more manageable size
        # This helps with memory issues and PDF generation
        max_dimension = 2000  # Maximum width or height
        if img.width > max_dimension or img.height > max_dimension:
            # Calculate new dimensions while preserving aspect ratio
            if img.width > img.height:
                new_width = max_dimension
                new_height = int(img.height * (max_dimension / img.width))
            else:
                new_height = max_dimension
                new_width = int(img.width * (max_dimension / img.height))
            
            # Resize the image
            img = img.resize((new_width, new_height), Image.LANCZOS)
            st.write(f"Resized image from {img.width}x{img.height} to {new_width}x{new_height}")
        
        # Check if image is empty or all white
        if img.mode == 'RGBA':
            # Convert RGBA to RGB with white background
            background = Image.new('RGB', img.size, (255, 255, 255))
            background.paste(img, mask=img.split()[3])  # 3 is the alpha channel
            img = background
        elif img.mode != 'RGB':
            img = img.convert('RGB')
        
        # Create PDF with proper dimensions
        width, height = img.size
        
        # Create PDF in a separate try block to isolate PDF creation issues
        try:
            pdf_buffer = BytesIO()
            pdf = canvas.Canvas(pdf_buffer)
            pdf.setPageSize((width, height))
            pdf.drawInlineImage(img, 0, 0, width=width, height=height)
            pdf.showPage()
            pdf.save()
            return pdf_buffer.getvalue()
        except Exception as pdf_error:
            st.error(f"Error creating PDF: {str(pdf_error)}")
            
            # Try an alternative approach for problematic images
            try:
                # Save image to a temporary file
                temp_img_path = f"temp_img_{datetime.now().strftime('%Y%m%d%H%M%S')}.jpg"
                img.save(temp_img_path, "JPEG")
                
                # Create PDF from the saved image
                pdf_buffer = BytesIO()
                pdf = canvas.Canvas(pdf_buffer)
                pdf.setPageSize((width, height))
                pdf.drawImage(temp_img_path, 0, 0, width=width, height=height)
                pdf.showPage()
                pdf.save()
                
                # Clean up temporary file
                if os.path.exists(temp_img_path):
                    os.remove(temp_img_path)
                    
                return pdf_buffer.getvalue()
            except Exception as alt_error:
                st.error(f"Alternative PDF creation also failed: {str(alt_error)}")
                return None
            
    except requests.exceptions.RequestException as e:
        st.error(f"Network error accessing URL {image_url}: {str(e)}")
        return None
    except PIL.UnidentifiedImageError:
        st.error(f"Could not identify image format for URL: {image_url}")
        # Try to save the raw content for debugging
        try:
            with open(f"debug_image_{datetime.now().strftime('%Y%m%d%H%M%S')}.bin", "wb") as f:
                f.write(response.content)
            st.error("Saved raw content for debugging")
        except:
            pass
        return None
    except AttributeError as e:
        if "'NoneType' object has no attribute 'seek'" in str(e):
            st.error(f"Invalid or empty image data from URL: {image_url}")
        else:
            st.error(f"Attribute error processing image {image_url}: {str(e)}")
        return None
    except MemoryError:
        st.error(f"Memory error processing large image from URL: {image_url}")
        return None
    except Exception as e:
        st.error(f"Error processing image {image_url}: {str(e)}")
        return None

def process_file(uploaded_file):
    output_base = get_output_filename(uploaded_file.name)
    df = pd.read_csv(uploaded_file)
    total_rows = len(df)
    
    if not os.path.exists("images"):
        os.makedirs("images")

    df["pdf_index"] = range(total_rows)
    df["date"] = df["date"].apply(format_date)
    df["receipt_status"] = "Available"

    # Convert boolean to Yes/No for company_credit_card
    if "company_credit_card" in df.columns:
        df["company_credit_card"] = df["company_credit_card"].map({True: "Yes", False: "No"})

    columns = [
        "pdf_index", "id", "date", "customer", "user", "picture",
        "type", "description", "amount", "company_credit_card", "receipt_status"
    ]
    df = df[columns]

    if "picture" in df.columns:
        df.loc[df["picture"].isna(), "receipt_status"] = "Missing"
        df.loc[df["picture"].str.len() == 0, "receipt_status"] = "Missing"

    excel_path = create_excel_file(df, output_base)
    zip_path = f"{output_base}.zip"

    progress_bar = st.progress(0)
    status_text = st.empty()
    
    # Add a container for error messages
    error_container = st.container()
    problem_images = []

    # Process PDFs and store them in memory
    pdf_files = {}
    if "picture" in df.columns:
        for index, row in df.iterrows():
            status_text.text(f"Processing receipt {index + 1} of {total_rows}")
            progress_bar.progress((index + 1) / total_rows)
            
            if row["receipt_status"] == "Available":
                pdf_content = create_pdf_from_image(row.get("picture"))
                if not pdf_content:
                    df.at[index, "receipt_status"] = "Failed"
                    problem_images.append(f"Row {index+1}: {row['id']} - {row['date']} - Failed to process")
                else:
                    pdf_name = f"{index:04d}_{row['date']}_{row['id']}.pdf"
                    pdf_files[pdf_name] = pdf_content

    # Display any problem images
    if problem_images:
        with error_container:
            st.warning("Some images could not be processed correctly:")
            for msg in problem_images:
                st.write(msg)

    # Create final Excel file with updated statuses
    excel_path = create_excel_file(df, output_base)

    # Create zip file with Excel and PDFs
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
        zipf.write(excel_path, f"{output_base}.xlsx")
        
        for pdf_name, pdf_content in pdf_files.items():
            zipf.writestr(f"receipts/{pdf_name}", pdf_content)

    # Cleanup temporary Excel file
    if os.path.exists(excel_path):
        os.remove(excel_path)

    # Add to processing history
    history_data = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "filename": uploaded_file.name,
        "processed_file": zip_path,
        "total_receipts": total_rows,
        "processed_receipts": sum(df["receipt_status"] == "Available")
    }
    add_to_history(history_data)

    status_text.text("Processing complete!")
    return zip_path

def delete_from_history(timestamp):
    conn = sqlite3.connect('expense_processor.db')
    c = conn.cursor()
    
    # Get file path before deletion
    c.execute('SELECT processed_file FROM processing_history WHERE timestamp = ?', (timestamp,))
    result = c.fetchone()
    if result:
        file_path = result[0]
        # Delete the zip file if it exists
        if os.path.exists(file_path):
            os.remove(file_path)
    
    # Delete from database
    c.execute('DELETE FROM processing_history WHERE timestamp = ?', (timestamp,))
    conn.commit()
    conn.close()

def test_image_url(url):
    """Test function to diagnose image URL access issues"""
    st.subheader("Image URL Test")
    
    # Try with different headers
    headers_options = [
        {},  # No headers
        {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        },
        {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Accept': 'image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8',
            'Referer': 'https://spesr.natron.io/',
        }
    ]
    
    for i, headers in enumerate(headers_options):
        st.write(f"Attempt {i+1} with headers: {headers}")
        try:
            response = requests.get(url, headers=headers, timeout=15, allow_redirects=True)
            st.write(f"Status code: {response.status_code}")
            st.write(f"Content type: {response.headers.get('Content-Type')}")
            st.write(f"Content length: {len(response.content)}")
            
            if response.status_code == 200 and response.content:
                st.success("Successfully retrieved content")
                try:
                    img = Image.open(BytesIO(response.content))
                    st.write(f"Image mode: {img.mode}, Size: {img.size}")
                    st.image(img, caption="Retrieved image")
                    st.success("Successfully parsed as image")
                except Exception as e:
                    st.error(f"Failed to parse as image: {str(e)}")
            else:
                st.error("Failed to retrieve valid content")
        except Exception as e:
            st.error(f"Request failed: {str(e)}")

def main():
    init_database()
    
    st.title("🧾 Expense Report Processor")
    
    # Add tabs for main functionality and diagnostics
    tab1, tab2 = st.tabs(["Process Reports", "Diagnostics"])
    
    with tab1:
        col1, col2 = st.columns([2, 1])
        
        with col1:
            st.markdown("""
            ### Upload Expense Report
            Upload your expense report CSV file (format: expenses_YYYY-M.csv)
            """)
            
            uploaded_file = st.file_uploader(
                label="Upload CSV file",
                type="csv",
                label_visibility="collapsed"
            )

            if uploaded_file:
                if not uploaded_file.name.startswith("expenses_"):
                    st.warning("File name should follow the format: expenses_YYYY-M.csv")
                elif st.button("Process Report", type="primary"):
                    with st.spinner("Processing expense report..."):
                        zip_path = process_file(uploaded_file)
                        with open(zip_path, "rb") as file:
                            st.success("Processing complete! Click below to download.")
                            st.download_button(
                                label="📥 Download Processed Files",
                                data=file,
                                file_name=os.path.basename(zip_path),
                                mime="application/zip",
                                key="main_download"
                            )

        with col2:
            st.markdown("### Processing History")
            history = get_processing_history()
            if history:
                for idx, item in enumerate(history):
                    with st.expander(f"📋 {item['filename']} ({item['timestamp']})"):
                        col1, col2 = st.columns([3, 1])
                        
                        with col1:
                            st.write(f"Total receipts: {item['total_receipts']}")
                            st.write(f"Processed receipts: {item['processed_receipts']}")
                            if os.path.exists(item['processed_file']):
                                st.download_button(
                                    label="📥 Download",
                                    data=open(item['processed_file'], "rb"),
                                    file_name=os.path.basename(item['processed_file']),
                                    mime="application/zip",
                                    key=f"history_download_{idx}"
                                )
                        
                        with col2:
                            if st.button("🗑️ Delete", key=f"delete_{idx}"):
                                delete_from_history(item['timestamp'])
                                st.rerun()
            else:
                st.info("No files processed yet")
    
    with tab2:
        st.subheader("Image URL Diagnostics")
        test_url = st.text_input("Enter image URL to test")
        if test_url and st.button("Test URL"):
            test_image_url(test_url)

if __name__ == "__main__":
    main()